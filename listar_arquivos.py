from pathlib import Path
from openpyxl import Workbook

# Pasta X dentro de Y dentro de Z
pasta = Path.home() / "Z" / "Y" / "X"

# Criar planilha
wb = Workbook()
ws = wb.active
ws.title = "Arquivos"

# Cabeçalho
ws["A1"] = "Nome do Arquivo"

# Adicionar arquivos à planilha
for linha, arquivo in enumerate(pasta.iterdir(), start=2):
    if arquivo.is_file():
        ws.cell(row=linha, column=1, value=arquivo.name)

# Salvar o Excel na própria pasta Y
arquivo_excel = Path.home() / "Downloads" / "lista_arquivos.xlsx"
wb.save(arquivo_excel)

print(f"Planilha salva em: {arquivo_excel}")